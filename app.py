import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

FILA_START    = 9
FILA_FAMILIA  = 7
FILA_STOCK_ID = 3

def init_page():
    st.set_page_config(layout="wide")
    st.title("Detector de Saltos en Series Históricas")

def initSettings():
    with st.expander("Parámetros de la búsqueda", expanded=True):
        umbral_rf          = st.number_input("Umbral Renta Fija (%)", min_value=0.0, max_value=100.0, value=4.0) / 100
        umbral_rv          = st.number_input("Umbral Renta Variable (%)", min_value=0.0, max_value=100.0, value=8.0) / 100
        dias_hueco         = st.number_input("Días consecutivos para detectar huecos", min_value=1, max_value=30, value=2)
        dias_repetidos_max = st.number_input("Días consecutivos al final de la serie para ver si está desactualizado.", min_value=1, max_value=30, value=10)
        if st.button("Comenzar la búsqueda", type="primary"):
            return True, dias_repetidos_max, umbral_rf, umbral_rv, dias_hueco
        else:
            return False, None, None, None, None


def load_file():
    uploaded_file = st.file_uploader("Sube un archivo Excel", type=["xlsx", "xlsm"])
    return uploaded_file

def backup_file(uploaded_file):
    temp_excel_path = "archivo_original.xlsx"
    with open(temp_excel_path, "wb") as f:
        f.write(uploaded_file.getbuffer())

    return temp_excel_path

def get_tipos_series(df):
    tipos_series = []
    columnas_validas = []
    familias_rf = [1, 2, 3, 4, 5, 6, 7, 8]  # tu lista de familias RF

    for idx, col in enumerate(df.columns[2:]):
        valor_fila_stk_id = df.iloc[FILA_STOCK_ID - 1, idx + 2]

        if pd.isna(valor_fila_stk_id):
            continue  # salta columnas sin número en fila 3

        columnas_validas.append(col)

        valor_fila_fam = df.iloc[FILA_FAMILIA - 1, idx + 2]
        if pd.isna(valor_fila_fam):
            tipos_series.append('RV')
        else:
            tipos_series.append('RF' if int(valor_fila_fam) in familias_rf else 'RV')

    return columnas_validas, tipos_series

def check_file_ok(temp_excel_path):
    df_temp = pd.read_excel(temp_excel_path, engine="openpyxl", header=None)
    # tipos_series = df_temp.iloc[3]
    columnas_validas, tipos_series = get_tipos_series(df_temp)
    print('tipos_series', tipos_series)
    # Añadir también las dos primeras columnas (fecha y etiqueta u otras)
    columnas_finales = list(df_temp.columns[:2]) + columnas_validas

    # column_names = df_temp.iloc[2]
    column_names = df_temp.iloc[2, columnas_finales]
    print('column_names', column_names)
    # df_raw = df_temp.iloc[(FILA_START - 1):]
    df_raw = df_temp.iloc[(FILA_START - 1):, columnas_finales]
    df_raw = df_raw[::-1].dropna(how='all').iloc[::-1]
    df_raw.columns = column_names
    df_raw.reset_index(drop=True, inplace=True)

    if df_raw.empty or df_raw.shape[1] < 3:
        st.error("El archivo debe tener al menos una columna de fechas (columna 2) y dos de datos (desde la columna 3).")
        return False, None, None, columnas_validas
    else:
        st.success("Archivo cargado correctamente.")
        return True, df_raw, tipos_series, columnas_validas

# Devuelve el DataFrame indexado por la fecha columna. La convertimos a datetime por si acaso.
def get_ordered_dataframe_by_date(df_raw):
    df = df_raw.copy()
    fecha_col = df.columns[1]
    df[fecha_col] = pd.to_datetime(df[fecha_col])
    df = df.sort_values(by=fecha_col).set_index(fecha_col)

    return df

def check_desactualizada(serie, dias_repetidos_max):
    ult_val = serie.iloc[-1]
    ultimos_repetidos = (serie[::-1] == ult_val).cumprod().sum()
    if ultimos_repetidos >= dias_repetidos_max:
        return True, ultimos_repetidos, ult_val
    else:
        return False, None, None

def get_celdas_desactualizadas_serie(df_raw, serie, ultimos_repetidos, idx):
    try:
        celdas_desactualizadas = []
        fecha_col              = df_raw.columns[1]
        fila_inicio            = df_raw[df_raw[fecha_col] == serie.index[-ultimos_repetidos]].index[0] + FILA_START

        for i in range(ultimos_repetidos):
            fila_excel = fila_inicio + i
            col_excel  = get_column_letter(idx + 4)
            # print((col_excel, fila_excel, "FF9999FF"))
            celdas_desactualizadas.append((col_excel, fila_excel, "FF9999FF"))  # azul

        return celdas_desactualizadas
    except IndexError:
        # print('Error', IndexError)
        return []
    
def check_huecos(df_raw, serie, idx, col, dias_hueco):
    dias_huecos   = []
    celdas_huecos = []
    fecha_col     = df_raw.columns[1]

    rep_count = 1
    start_idx = None
    for i in range(1, len(serie)):
        if pd.isna(serie.iloc[i]) or pd.isna(serie.iloc[i - 1]):
            rep_count = 1
            start_idx = None
            continue
        if serie.iloc[i] == serie.iloc[i - 1]:
            if rep_count == 1:
                start_idx = i - 1
            rep_count += 1
        else:
            if rep_count >= dias_hueco:
                valor_rep = serie.iloc[start_idx]
                fecha_inicio = serie.index[start_idx]
                fecha_fin = serie.index[i - 1]
                dias_huecos.append({
                    "Serie": col,
                    "Tipo": "Hueco",
                    "Valor": valor_rep,
                    "Fecha inicio": fecha_inicio.strftime("%Y-%m-%d") if pd.notna(fecha_inicio) else "",
                    "Fecha fin": fecha_fin.strftime("%Y-%m-%d") if pd.notna(fecha_fin) else "",
                    "Repeticiones": rep_count
                })
                for j in range(rep_count):
                    fecha_hueco = serie.index[i - j - 1]
                    fila_excel = df_raw[df_raw[fecha_col] == fecha_hueco].index[0] + FILA_START
                    col_excel = get_column_letter(idx + 4)
                    celdas_huecos.append((col_excel, fila_excel, "FFFFC000"))
            rep_count = 1
            start_idx = None
    if rep_count >= dias_hueco:
        valor_rep = serie.iloc[start_idx]
        fecha_inicio = serie.index[start_idx]
        fecha_fin = serie.index[-1]
        dias_huecos.append({
            "Serie": col,
            "Tipo": "Hueco",
            "Valor": valor_rep,
            "Fecha inicio": fecha_inicio.strftime("%Y-%m-%d") if pd.notna(fecha_inicio) else "",
            "Fecha fin": fecha_fin.strftime("%Y-%m-%d") if pd.notna(fecha_fin) else "",
            "Repeticiones": rep_count
        })
        for j in range(rep_count):
            fecha_hueco = serie.index[-j - 1]
            fila_excel = df_raw[df_raw[fecha_col] == fecha_hueco].index[0] + FILA_START
            col_excel = get_column_letter(idx + 4)
            celdas_huecos.append((col_excel, fila_excel, "FFFFC000"))

    return dias_huecos, celdas_huecos

def check_saltos_y_rebotes(df_raw, serie, idx, col, tipos_series, umbral_rf, umbral_rv):
    resultados_marcados = []
    rebotes_detectados  = []
    celdas_saltos       = []
    celdas_rebotes      = []
    # tipo                = tipos_series[col] if col in tipos_series.index else 'RV'
    tipo                = tipos_series[idx] if idx < len(tipos_series) else 'RV'
    umbral              = umbral_rf if tipo == 'RF' else umbral_rv
    fecha_col           = df_raw.columns[1]

    # print('SERIE LEN: ', len(serie))

    try:
        serie_pct = serie.pct_change()

        for i in range(1, len(serie_pct)):
            fecha  = serie_pct.index[i]
            valor  = serie.iloc[i]
            cambio = serie_pct.iloc[i]

            if abs(cambio) > umbral:
                resultados_marcados.append({
                    "Activo": col,
                    "Fecha": fecha,
                    "Valor": valor,
                    "Diferencia": cambio * 100
                })
                try:
                    fila_excel = df_raw[df_raw[fecha_col] == fecha].index[0] + FILA_START
                    col_excel  = get_column_letter(idx + 4)
                    celdas_saltos.append((col_excel, fila_excel, "FFFF6666"))  # rojo
                except IndexError:
                    print('Error salto')

            if i < len(serie_pct) - 1:
                cambio_sig = serie_pct.iloc[i + 1]
                if abs(cambio) > umbral and abs(cambio_sig) > umbral:
                    if np.sign(cambio) != np.sign(cambio_sig):
                        rebotes_detectados.append({
                            "Activo": col,
                            "Fecha": fecha,
                            "Valor": valor,
                            "Diferencia": cambio * 100
                        })
                        try:
                            fila_excel = df_raw[df_raw[fecha_col] == fecha].index[0] + FILA_START
                            col_excel = get_column_letter(idx + 4)
                            celdas_rebotes.append((col_excel, fila_excel, "FFFF6666"))  # rojo
                        except IndexError:
                            print('Error rebote')
        return resultados_marcados, rebotes_detectados, celdas_saltos, celdas_rebotes
    except ValueError:
        print('Error columna: ', col)

def process_serie_column(df_raw, df, idx, col, dias_repetidos_max, umbral_rf, umbral_rv, dias_hueco, tipos_series):
    info_desactualizada    = (None, None, None)
    celdas_desactualizadas = []
    dias_huecos            = []
    celdas_huecos          = []
    resultados_marcados    = []
    rebotes_detectados     = []
    celdas_saltos          = []
    celdas_rebotes         = []

    serie = df[col]
    serie = serie.loc[:serie.last_valid_index()]
    serie = pd.to_numeric(serie, errors="coerce")

    # Comprobar si está desactualizada la serie.
    esta_desactualizada, ultimos_repetidos, ult_val = check_desactualizada(serie, dias_repetidos_max)
    if esta_desactualizada:
        info_desactualizada    = (col, ult_val, ultimos_repetidos)
        celdas_desactualizadas = get_celdas_desactualizadas_serie(df_raw, serie, ultimos_repetidos, idx) 

    # Comprobar si tiene huecos la serie.
    dias_huecos, celdas_huecos = check_huecos(df_raw, serie, idx, col, dias_hueco)

    # Comprobar saltos y rebotes
    resultados_marcados, rebotes_detectados, celdas_saltos, celdas_rebotes = check_saltos_y_rebotes(df_raw, serie, idx, col, tipos_series, umbral_rf, umbral_rv)

    return info_desactualizada, celdas_desactualizadas, dias_huecos, celdas_huecos, resultados_marcados, rebotes_detectados, celdas_saltos, celdas_rebotes

def print_resumen(resultados_marcados, desactualizadas, dias_huecos, rebotes_detectados):
    resumen_errores = []

    for resultado in resultados_marcados:
        resumen_errores.append({
            "Serie": resultado["Activo"],
            "Tipo": "Salto",
            "Valor": resultado["Valor"],
            "Fecha": resultado["Fecha"].strftime("%Y-%m-%d") if pd.notna(resultado["Fecha"]) else "",
            "Cambio (%)": round(resultado["Diferencia"], 2)
        })
    for col, val, rep in desactualizadas:
        resumen_errores.append({
            "Serie": col,
            "Tipo": "Desactualizada",
            "Valor": val,
            "Días repetidos": rep
        })
    for h in dias_huecos:
        resumen_errores.append(h)
    for r in rebotes_detectados:
        resumen_errores.append({
            "Serie": r["Activo"],
            "Tipo": "Rebote",
            "Fecha": r["Fecha"].strftime("%Y-%m-%d") if pd.notna(r["Fecha"]) else "",
            "Valor": r["Valor"],
            "Cambio (%)": round(r["Diferencia"], 2)
        })

    with st.expander('⚠️ Resumen de errores'):
        if desactualizadas:
            st.warning("⚠️ Series desactualizadas detectadas:")
            for col, val, rep in desactualizadas:
                st.write(f"- {col}: valor {val} repetido {rep} días consecutivos al final")

        if resumen_errores:
            st.subheader("📋 Resumen de errores detectados")
            resumen_df = pd.DataFrame(resumen_errores)
            st.dataframe(resumen_df, use_container_width=True)

    return resumen_errores

def print_cards(df, resumen_errores, resultados_marcados, rebotes_detectados, dias_hueco, desactualizadas, tipos_series, columnas_validas):
    print("Tipos series: ", tipos_series)
    
    for idx, col in enumerate(df.columns[2:]):
        incidencias = [e for e in resumen_errores if e.get("Serie") == col]
        estado = f"✅ OK" if not incidencias else f"⚠️ {len(incidencias)} incidencia(s)"
        print("idx,col: ", idx, " - ", col)
        tipo = tipos_series[idx + 1] if idx < len(tipos_series) else "RV"

        with st.expander(f"📈 {col} · {estado} · {tipo}"):

            if incidencias:
                st.info(f"🔍 {len(incidencias)} incidencias detectadas para esta serie.")
            else:
                st.info("Sin incidencias detectadas en esta serie.")

            serie = df[col].copy()
            serie = serie[serie.index.notna()]

            # Intentar convertir la serie a valores numéricos
            serie = pd.to_numeric(serie, errors="coerce")

            # Saltar si no hay al menos 2 valores numéricos válidos
            if serie.notna().sum() < 2:
                st.warning("No hay suficientes datos numéricos para graficar esta serie.")
                continue

            fig, ax = plt.subplots()
            ax.plot(serie.index, serie.values, label=col, color="steelblue")

            # Detectar saltos y rebotes
            saltos_fechas = [r["Fecha"] for r in resultados_marcados if r["Activo"] == col]
            rebotes_fechas = [r["Fecha"] for r in rebotes_detectados if r["Activo"] == col]

            fechas_index = serie.index

            def fechas_a_indices(fechas, index):
                indices = []
                for f in fechas:
                    try:
                        pos = index.get_loc(pd.to_datetime(f))
                        indices.append(pos)
                    except KeyError:
                        pass
                return indices

            saltos_idx = fechas_a_indices(saltos_fechas, fechas_index)
            rebotes_idx = fechas_a_indices(rebotes_fechas, fechas_index)

            ax.scatter(serie.index[saltos_idx], serie.iloc[saltos_idx], color='red', label='Saltos', zorder=5)
            ax.scatter(serie.index[rebotes_idx], serie.iloc[rebotes_idx], color='orange', label='Rebotes', zorder=5)

            # Pintar tramos huecos (repetidos)
            rep_count = 1
            start_idx = None
            for i in range(1, len(serie)):
                if pd.isna(serie.iloc[i]) or pd.isna(serie.iloc[i-1]):
                    if rep_count >= dias_hueco and start_idx is not None:
                        ax.axvspan(serie.index[start_idx], serie.index[i - 1], color='orange', alpha=0.2)
                    rep_count = 1
                    start_idx = None
                    continue
                if serie.iloc[i] == serie.iloc[i - 1]:
                    if rep_count == 1:
                        start_idx = i - 1
                    rep_count += 1
                else:
                    if rep_count >= dias_hueco and start_idx is not None:
                        ax.axvspan(serie.index[start_idx], serie.index[i - 1], color='orange', alpha=0.2)
                    rep_count = 1
                    start_idx = None
            if rep_count >= dias_hueco and start_idx is not None:
                ax.axvspan(serie.index[start_idx], serie.index[len(serie) - 1], color='orange', alpha=0.2)

            # Pintar tramos desactualizados
            for col_des, val, rep in desactualizadas:
                if col_des == col and rep is not None and rep <= len(serie):
                    ax.axvspan(serie.index[-rep], serie.index[-1], color='blue', alpha=0.2)

            ax.set_title(f"Evolución de {col}")
            ax.set_ylabel("Valor")
            ax.legend()
            ax.grid(True)
            st.pyplot(fig)

            if incidencias:
                st.write("**Incidencias:**")
                st.dataframe(pd.DataFrame(incidencias), use_container_width=True)

def generar_mapa_errores(df, resumen_errores, resultados_marcados, rebotes_detectados, desactualizadas, dias_hueco):
    mapa_errores = {}

    for col in df.columns[2:]:
        errores = []

        # H: Hueco (datos repetidos)
        serie = pd.to_numeric(df[col], errors="coerce")
        rep_count = 1
        for i in range(1, len(serie)):
            if pd.isna(serie.iloc[i]) or pd.isna(serie.iloc[i - 1]):
                rep_count = 1
                continue
            if serie.iloc[i] == serie.iloc[i - 1]:
                rep_count += 1
                if rep_count >= dias_hueco:
                    errores.append("H")
                    break
            else:
                rep_count = 1

        # S: Saltos
        if any(r["Activo"] == col for r in resultados_marcados):
            errores.append("S")

        # R: Rebotes
        if any(r["Activo"] == col for r in rebotes_detectados):
            errores.append("R")

        # D: Desactualizada
        if any(d[0] == col for d in desactualizadas):
            errores.append("D")

        mapa_errores[col] = "-".join(sorted(set(errores))) if errores else ""
    return mapa_errores


def set_new_excel_to_download(temp_excel_path, celdas_sospechosas, resumen_errores, mapa_errores):
    wb = load_workbook(temp_excel_path)
    ws = wb.active
    for col, fila, color in celdas_sospechosas:
        ws[f"{col}{fila}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Convertimos el generador en lista
    columnas_fila_3 = list(ws.iter_cols(min_row=3, max_row=3, values_only=True))

    # Recorremos cada columna de la fila 3
    for col_idx, columna in enumerate(columnas_fila_3, start=1):
        col_name = columna[0]  # Solo hay una fila, por lo que tomamos el primer valor
        siglas = mapa_errores.get(col_name, "")
        if siglas:
            ws.cell(row=2, column=col_idx, value=siglas)
            ws.cell(row=2, column=col_idx).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    if resumen_errores:
        ws_resumen = wb.create_sheet("Resumen_Errores")
        resumen_df = pd.DataFrame(resumen_errores)
        for i, col_name in enumerate(resumen_df.columns, start=1):
            ws_resumen.cell(row=1, column=i).value = col_name
        for row_idx, row in enumerate(resumen_df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row, start=1):
                ws_resumen.cell(row=row_idx, column=col_idx).value = val

    excel_anotado = "saltos_anotados.xlsx"
    wb.save(excel_anotado)

    with open(excel_anotado, "rb") as f:
        st.download_button("Descargar Excel con saltos señalados", f.read(),
                        file_name=excel_anotado,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def process_file(uploaded_file):
    if uploaded_file:
        temp_excel_path = backup_file(uploaded_file)
        ok, df_raw, tipos_series, columnas_validas = check_file_ok(temp_excel_path)

        if ok:
            comenzar, dias_repetidos_max, umbral_rf, umbral_rv, dias_hueco = initSettings()

            if comenzar:
                df = get_ordered_dataframe_by_date(df_raw)

                celdas_sospechosas  = []
                rebotes_detectados  = []
                resultados_marcados = []
                dias_huecos         = []
                desactualizadas     = []

                # Ahora nos vamos a recorrer las diferentes columnas, teniendo en cuenta
                # que empezamos desde la tercera columna.
                for idx, col in enumerate(df.columns[2:]):
                    info_desactualizada, celdas_desact_serie, dias_huecos_serie, celdas_huecos_serie, resultados_marcados_serie, rebotes_detectados_serie, celdas_saltos_serie, celdas_rebotes_serie = process_serie_column(df_raw, df, idx, col, dias_repetidos_max, umbral_rf, umbral_rv, dias_hueco, tipos_series)
                    
                    dias_huecos.extend(dias_huecos_serie)
                    celdas_sospechosas.extend(celdas_huecos_serie)

                    if info_desactualizada[0] != None:
                        desactualizadas.append(info_desactualizada)
                        celdas_sospechosas.extend(celdas_desact_serie)

                    resultados_marcados.extend(resultados_marcados_serie)
                    celdas_sospechosas.extend(celdas_saltos_serie)
                    
                    rebotes_detectados.extend(rebotes_detectados_serie)
                    celdas_sospechosas.extend(celdas_rebotes_serie)
                    
                # RESUMEN DE ERRORES
                resumen_errores = print_resumen(resultados_marcados, desactualizadas, dias_huecos, rebotes_detectados)

                mapa_errores = generar_mapa_errores(df, resumen_errores, resultados_marcados, rebotes_detectados, desactualizadas, dias_hueco)

                # Son las diferentes tarjetas
                print_cards(df, resumen_errores, resultados_marcados, rebotes_detectados, dias_hueco, desactualizadas, tipos_series, columnas_validas)

                # Prepara el nuevo Excel con los cambios para descargar
                set_new_excel_to_download(temp_excel_path, celdas_sospechosas, resumen_errores, mapa_errores)

# COMIENZA LA EJECUCIÓN DE LA APLICACIÓN
init_page()
process_file(uploaded_file=load_file())





